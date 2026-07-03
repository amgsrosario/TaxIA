# -*- coding: utf-8 -*-
"""
Etapa 9B.1 — Conversão do Excel do piloto para CSV técnico.

Lê pilot/input/TaxIA_20_casos_piloto_v01.xlsx (folha "Template TaxIA") SEM o
alterar e produz:
  - pilot/working/taxia-pilot-main.csv     (casos principais, contrato real)
  - pilot/working/taxia-pilot-reserve.csv  (apenas o caso RESERVA)
  - dados de inspecção/validação em JSON no stdout para o relatório

Política:
  - "PENDENTE DE CONFIRMAÇÃO" em validFrom/validTo/sourceUrl → campo técnico
    vazio, pendência copiada para notes, aviso no relatório, nunca inventar.
  - Datas normalizadas para YYYY-MM-DD.
  - Colunas editoriais (sourceType, sourceTitle, sourceUrl, decision) ficam
    FORA do CSV técnico — apenas nos relatórios.
  - Mapeamento documentado (não silencioso) dos nomes de tema para os enums
    reais do backend; valores não mapeáveis seguem como estão e são
    assinalados como enum inválido.
  - Conteúdo fiscal (perguntas, respostas, artigos, taxas, prazos) NUNCA é
    alterado.
"""
import csv
import io
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "pilot/input/TaxIA_20_casos_piloto_v01.xlsx"
MAIN_CSV = ROOT / "pilot/working/taxia-pilot-main.csv"
RESERVE_CSV = ROOT / "pilot/working/taxia-pilot-reserve.csv"

PENDING = "PENDENTE DE CONFIRMAÇÃO"

# Contrato real (KnowledgeQuestionAnswerImportService.parseCsv / ImportRow)
CONTRACT_COLUMNS = [
    "externalKey", "question", "answer", "topic", "subtopic", "jurisdiction",
    "riskLevel", "requiresHumanValidation", "sourceReference",
    "validFrom", "validTo", "notes",
]

# Enums reais do backend
TOPIC_ENUM = {"IVA", "IRC", "IRS", "SEGURANCA_SOCIAL", "TRABALHO",
              "CONTABILIDADE", "PROCEDIMENTO_TRIBUTARIO", "FATURACAO", "OUTROS"}
RISK_ENUM = {"LOW", "MEDIUM", "HIGH", "CRITICAL"}
SOURCE_TYPE_ENUM = {"LEGISLATION", "ADMINISTRATIVE_GUIDANCE", "CASE_LAW",
                    "OFFICIAL_FAQ", "INTERNAL_OPINION", "ACCOUNTING_STANDARD", "OTHER"}

# Mapeamento documentado nome-de-apresentação → enum (relatório de conversão)
TOPIC_MAP = {
    "IVA": "IVA",
    "IRC": "IRC",
    "IRS": "IRS",
    "Segurança Social": "SEGURANCA_SOCIAL",
    "Direito do Trabalho": "TRABALHO",
    "Contabilidade": "CONTABILIDADE",
    "Procedimento tributário": "PROCEDIMENTO_TRIBUTARIO",
    "Faturação": "FATURACAO",
}

EXTERNAL_KEY_MAX = 240  # EXTERNAL_KEY_MAX_IMPORT_LENGTH


def cell_str(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def is_pending(value: str) -> bool:
    return value.strip().upper() == PENDING.upper() if value else False


def normalize_date(raw: str):
    """→ (valor_normalizado_ou_vazio, erro_ou_None)."""
    if not raw:
        return "", None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d"), None
        except ValueError:
            continue
    return "", f"data inválida: {raw!r}"


def main():
    wb = load_workbook(XLSX, data_only=False, read_only=False)
    assert wb.sheetnames == ["Registo 20 casos", "Template TaxIA", "Pendências", "Balanço"], wb.sheetnames

    ws = wb["Template TaxIA"]
    headers = [cell_str(c.value) for c in ws[3]]

    inspection = {
        "sheets": wb.sheetnames,
        "template_headers": headers,
        "formulas": [],
        "empty_cells": [],
        "newlines_in_cells": [],
        "date_cell_types": [],
        "boolean_values": set(),
    }

    # Inventário técnico de todas as folhas (fórmulas, tipos)
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    inspection["formulas"].append(f"{name}!{c.coordinate}")

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=4), start=4):
        values = [c.value for c in row]
        if not any(v is not None and str(v).strip() for v in values):
            continue
        rec = dict(zip(headers, values))
        rec["_excel_row"] = i
        # tipos de data
        for f in ("validFrom", "validTo"):
            idx = headers.index(f)
            v = values[idx]
            inspection["date_cell_types"].append(
                {"row": i, "field": f, "type": type(v).__name__, "value": cell_str(v)})
        inspection["boolean_values"].add(cell_str(rec.get("humanReviewRequired")))
        rows.append(rec)

    main_rows, reserve_rows, report_rows = [], [], []

    for rec in rows:
        row_report = {
            "excel_row": rec["_excel_row"],
            "externalKey": cell_str(rec.get("externalKey")),
            "sourceTitle": cell_str(rec.get("sourceTitle")),
            "decision": cell_str(rec.get("decision")),
            "sourceType_excel": cell_str(rec.get("sourceType")),
            "sourceUrl_excel": cell_str(rec.get("sourceUrl")),
            "warnings": [],
            "errors": [],
            "pending": [],
        }

        topic_raw = cell_str(rec.get("topic"))
        topic = TOPIC_MAP.get(topic_raw, topic_raw)
        if topic.upper() not in TOPIC_ENUM:
            row_report["errors"].append(f"topic inválido: {topic_raw!r}")
        elif topic != topic_raw:
            row_report["warnings"].append(
                f"topic mapeado documentadamente: {topic_raw!r} → {topic}")

        risk = cell_str(rec.get("riskLevel")).upper()
        if risk not in RISK_ENUM:
            row_report["errors"].append(f"riskLevel inválido: {risk!r}")

        st = cell_str(rec.get("sourceType"))
        if st and st not in SOURCE_TYPE_ENUM:
            row_report["warnings"].append(
                f"sourceType editorial {st!r} não existe no enum do backend "
                "(campo NÃO faz parte do contrato de importação; fica só no relatório)")

        hrr = cell_str(rec.get("humanReviewRequired")).lower()
        if hrr not in ("true", "false"):
            row_report["errors"].append(f"humanReviewRequired inválido: {hrr!r}")

        notes = cell_str(rec.get("notes"))
        pending_notes = []

        vf_raw, vt_raw = cell_str(rec.get("validFrom")), cell_str(rec.get("validTo"))
        vf = vt = ""
        if is_pending(vf_raw):
            row_report["pending"].append("validFrom")
            pending_notes.append("validFrom PENDENTE DE CONFIRMAÇÃO")
        else:
            vf, err = normalize_date(vf_raw)
            if err:
                row_report["errors"].append(f"validFrom {err}")
        if is_pending(vt_raw):
            row_report["pending"].append("validTo")
            pending_notes.append("validTo PENDENTE DE CONFIRMAÇÃO")
        else:
            vt, err = normalize_date(vt_raw)
            if err:
                row_report["errors"].append(f"validTo {err}")

        url_raw = cell_str(rec.get("sourceUrl"))
        if is_pending(url_raw) or not url_raw:
            row_report["pending"].append("SOURCE_URL_PENDING")
            pending_notes.append("SOURCE_URL_PENDING — URL oficial por confirmar")
        elif not re.match(r"^https?://", url_raw):
            row_report["pending"].append("SOURCE_URL_PENDING")
            row_report["warnings"].append(f"sourceUrl inválida (não usada): {url_raw!r}")
            pending_notes.append("SOURCE_URL_PENDING — URL registada é inválida")

        if pending_notes:
            suffix = "[PENDÊNCIAS 9B.1: " + "; ".join(pending_notes) + "]"
            notes = f"{notes} {suffix}".strip() if notes else suffix
            if hrr != "true":
                # Política 9B.1 §6: casos com pendências mantêm revisão humana
                # obrigatória. Alteração explícita e documentada, nunca silenciosa.
                row_report["warnings"].append(
                    "requiresHumanValidation forçado para true (política de pendências "
                    "9B.1 §6) — o Excel tinha false")
                hrr = "true"

        technical = {
            "externalKey": cell_str(rec.get("externalKey")),
            "question": cell_str(rec.get("question")),
            "answer": cell_str(rec.get("answer")),
            "topic": topic,
            "subtopic": cell_str(rec.get("subtopic")),
            "jurisdiction": "PT",
            "riskLevel": risk,
            "requiresHumanValidation": hrr,
            "sourceReference": cell_str(rec.get("sourceReference")),
            "validFrom": vf,
            "validTo": vt,
            "notes": notes,
        }

        if not technical["question"]:
            row_report["errors"].append("pergunta vazia")
        if not technical["answer"]:
            row_report["errors"].append("resposta vazia")
        if len(technical["externalKey"]) > EXTERNAL_KEY_MAX:
            row_report["errors"].append("externalKey acima de 240 caracteres")
        if not technical["sourceReference"]:
            row_report["warnings"].append("caso sem sourceReference")
        if "\n" in technical["question"] or "\n" in technical["answer"]:
            inspection["newlines_in_cells"].append(rec["_excel_row"])

        row_report["technical"] = technical
        report_rows.append(row_report)

        if row_report["decision"].strip().upper() == "RESERVA":
            reserve_rows.append(technical)
        else:
            main_rows.append(technical)

    # Duplicados de externalKey no lote principal
    seen = {}
    for t in main_rows:
        seen.setdefault(t["externalKey"], []).append(t)
    duplicates = [k for k, v in seen.items() if len(v) > 1]

    MAIN_CSV.parent.mkdir(parents=True, exist_ok=True)
    RESERVE_CSV.parent.mkdir(parents=True, exist_ok=True)
    for path, data in ((MAIN_CSV, main_rows), (RESERVE_CSV, reserve_rows)):
        # UTF-8 SEM BOM: o parser do backend lê UTF-8 e não remove BOM do cabeçalho.
        with io.open(path, "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=CONTRACT_COLUMNS, quoting=csv.QUOTE_ALL)
            w.writeheader()
            for t in data:
                w.writerow(t)

    inspection["boolean_values"] = sorted(inspection["boolean_values"])
    out = {
        "inspection": inspection,
        "total_cases": len(rows),
        "main_cases": len(main_rows),
        "reserve_cases": len(reserve_rows),
        "duplicates_externalKey": duplicates,
        "rows": report_rows,
    }
    sys.stdout.reconfigure(encoding="utf-8")
    print(json.dumps(out, ensure_ascii=False, indent=1, default=str))


if __name__ == "__main__":
    main()
